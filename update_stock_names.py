#!/usr/bin/env python3
"""
사명 변경 자동 감지 및 Excel 업데이트 스크립트
GitHub Actions에서 매주 월요일 자동 실행됩니다.

동작 방식:
  1. FDR(FinanceDataReader)로 현재 KRX 전체 종목 이름→코드 조회
  2. DART corpCode.xml로 코드→현재 공식 이름 조회
  3. Excel 파일의 종목명과 비교해 사명 변경 감지
  4. 변경된 종목명을 Excel 파일에 일괄 반영
  5. 변경 내역을 stock_code_map.json에 저장 (다음 실행 시 이전 사명도 추적 가능)
"""

import os
import json
import zipfile
import argparse
import requests
import pandas as pd
import FinanceDataReader as fdr
from io import BytesIO
import xml.etree.ElementTree as ET
from datetime import datetime

# ── 설정 ──────────────────────────────────────────────────────
DART_API_KEY = os.environ.get('DART_API_KEY', '')

# 이름→코드 누적 매핑 파일 (구 사명도 계속 추적하기 위해 repo에 저장)
CODE_MAP_FILE = 'stock_code_map.json'

# 구 사명→현재 사명 누적 이력 (앱 검색에서 구 사명으로도 찾을 수 있도록)
ALIASES_FILE = 'name_aliases.json'

# 업데이트 대상 Excel 파일 목록
EXCEL_FILES = [
    '종목정리_종목순 정렬.xlsx',
    '시그널뷰_기업개요.xlsx',
    '시그널뷰_종목정리_핵심정리 및 테마.xlsx',
    '시그널뷰_테마별 기업개요.xlsx',
    '시그널뷰_관련테마.xlsx',
    'stock_combined.xlsx',
]

# 종목명/종목코드 컬럼 후보 (공백 제거 후 비교)
NAME_COLS = ['종목명', '종목이름', '종목']
CODE_COLS = ['종목코드', '단축코드', '코드', 'Code', 'code', 'StockCode', 'stock_code']


def normalize_stock_code(value) -> str:
    """엑셀/FDR/DART 종목코드를 비교 가능한 문자열로 정규화"""
    if pd.isna(value):
        return ''

    if isinstance(value, float):
        if value.is_integer():
            value = int(value)
        else:
            value = str(value)

    code = str(value).strip()
    if not code or code.lower() in ('nan', 'none', 'nat'):
        return ''

    if code.endswith('.0') and code[:-2].isdigit():
        code = code[:-2]

    code = code.replace("'", "").replace('"', "").strip().upper()
    if code.startswith('A') and len(code) == 7 and code[1:].isdigit():
        code = code[1:]
    if code.isdigit():
        code = code.zfill(6)
    return code


def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    """컬럼명 공백 제거 및 종목명/종목코드 표준화"""
    df.columns = df.columns.str.replace(' ', '').str.strip()
    rename_map = {
        **{col: '종목명' for col in NAME_COLS if col != '종목명'},
        **{col: '종목코드' for col in CODE_COLS if col != '종목코드'},
    }
    df.rename(columns=rename_map, inplace=True)
    if '종목코드' in df.columns:
        df['종목코드'] = df['종목코드'].apply(normalize_stock_code)
    return df


# ── 데이터 수집 ────────────────────────────────────────────────

def get_fdr_maps() -> tuple[dict, dict]:
    """FDR로 현재 KRX 전체 종목의 이름→코드, 코드→이름 딕셔너리 반환"""
    name_to_code = {}
    code_to_name = {}
    try:
        df = fdr.StockListing('KRX')
        for _, row in df.iterrows():
            code = normalize_stock_code(row.get('Code', ''))
            name = str(row.get('Name', '')).strip()
            if code and name:
                name_to_code[name] = code
                code_to_name[code] = name
    except Exception as e:
        print(f"  [FDR] KRX 조회 실패: {e}")
    return name_to_code, code_to_name


def get_fdr_name_to_code() -> dict:
    """기존 호출 호환용: FDR 이름→코드 딕셔너리 반환"""
    name_to_code, _ = get_fdr_maps()
    return name_to_code


def get_dart_code_to_name() -> dict:
    """DART corpCode.xml에서 종목코드→현재 회사명 딕셔너리 반환"""
    if not DART_API_KEY:
        print("  [DART] API 키 없음 — 스킵")
        return {}
    try:
        url = f"https://opendart.fss.or.kr/api/corpCode.xml?crtfc_key={DART_API_KEY}"
        resp = requests.get(url, timeout=60)
        resp.raise_for_status()

        with zipfile.ZipFile(BytesIO(resp.content)) as z:
            with z.open('CORPCODE.xml') as f:
                root = ET.parse(f).getroot()

        result = {}
        for item in root.findall('list'):
            stock_code = normalize_stock_code(item.find('stock_code').text or '')
            corp_name  = (item.find('corp_name').text  or '').strip()
            if stock_code:
                result[stock_code] = corp_name

        print(f"  [DART] {len(result)}개 법인 로드 완료")
        return result
    except Exception as e:
        print(f"  [DART] 오류: {e}")
        return {}


# ── 매핑 파일 관리 ─────────────────────────────────────────────

def load_code_map() -> dict:
    """저장된 이름→코드 누적 매핑 로드"""
    if os.path.exists(CODE_MAP_FILE):
        with open(CODE_MAP_FILE, encoding='utf-8') as f:
            loaded = json.load(f)
            return {str(name).strip(): normalize_stock_code(code) for name, code in loaded.items() if str(name).strip()}
    return {}


def save_code_map(mapping: dict):
    """이름→코드 누적 매핑 저장"""
    with open(CODE_MAP_FILE, 'w', encoding='utf-8') as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)


def load_aliases() -> dict:
    """구 사명→현재 사명 누적 이력 로드"""
    if os.path.exists(ALIASES_FILE):
        with open(ALIASES_FILE, encoding='utf-8') as f:
            return json.load(f)
    return {}


def save_aliases(aliases: dict):
    """구 사명→현재 사명 누적 이력 저장"""
    with open(ALIASES_FILE, 'w', encoding='utf-8') as f:
        json.dump(aliases, f, ensure_ascii=False, indent=2)


def resolve_alias_name(name: str, aliases: dict) -> str:
    """A→B→C처럼 여러 번 바뀐 사명을 최종 사명으로 정리"""
    current = str(name).strip()
    seen = set()
    while current in aliases and current not in seen:
        seen.add(current)
        next_name = str(aliases[current]).strip()
        if not next_name or next_name == current:
            break
        current = next_name
    return current


def resolve_code_by_name(name: str, name_to_code: dict, aliases: dict | None = None) -> str:
    """종목명 또는 구 사명으로 종목코드 찾기"""
    name = str(name).strip()
    if not name or name.lower() in ('nan', 'none', 'nat'):
        return ''

    code = normalize_stock_code(name_to_code.get(name, ''))
    if code:
        return code

    aliases = aliases or {}
    current_name = resolve_alias_name(name, aliases) if name in aliases else name
    if current_name != name:
        return normalize_stock_code(name_to_code.get(current_name, ''))

    return ''


def sync_aliases_from_duplicate_codes(
    code_map: dict,
    current_code_to_name: dict,
    aliases: dict,
) -> int:
    """
    stock_code_map.json에 같은 종목코드로 여러 이름이 있으면
    FDR 현재명을 기준으로 나머지 이름을 alias로 누적.
    """
    if not current_code_to_name:
        return 0

    code_to_names = {}
    for name, code in code_map.items():
        name = str(name).strip()
        code = normalize_stock_code(code)
        if not name or not code:
            continue
        code_to_names.setdefault(code, set()).add(name)

    changed = 0
    for code, names in sorted(code_to_names.items()):
        if len(names) < 2:
            continue

        current_name = str(current_code_to_name.get(code, '')).strip()
        if not current_name or current_name not in names:
            continue

        for old_name in sorted(names):
            if old_name == current_name:
                continue
            if aliases.get(old_name) != current_name:
                aliases[old_name] = current_name
                changed += 1

    return changed


# ── Excel 처리 ─────────────────────────────────────────────────

def get_all_stock_records() -> list[dict]:
    """모든 Excel 파일의 전체 시트에서 종목명/종목코드 수집"""
    records = []
    seen = set()
    for path in EXCEL_FILES:
        if not os.path.exists(path):
            continue
        try:
            xl = pd.ExcelFile(path, engine='openpyxl')
            for sheet in xl.sheet_names:
                df = xl.parse(sheet)
                df = clean_columns(df)
                name_col = '종목명' if '종목명' in df.columns else None
                code_col = '종목코드' if '종목코드' in df.columns else None

                if not name_col and not code_col:
                    continue

                for _, row in df.iterrows():
                    name = str(row.get(name_col, '')).strip() if name_col else ''
                    if name.lower() in ('nan', 'none', 'nat'):
                        name = ''
                    code = normalize_stock_code(row.get(code_col, '')) if code_col else ''

                    if not name and not code:
                        continue

                    key = (code, name)
                    if key in seen:
                        continue
                    seen.add(key)
                    records.append({
                        'file': path,
                        'sheet': sheet,
                        'name': name,
                        'code': code,
                    })
        except Exception as e:
            print(f"  [Excel] {path} 읽기 실패: {e}")
    return records


def apply_name_changes(name_changes: dict) -> int:
    """모든 Excel 파일에 {구 사명: 신 사명} 일괄 치환. 수정된 행 수 반환."""
    total = 0
    for path in EXCEL_FILES:
        if not os.path.exists(path):
            continue
        try:
            xl = pd.ExcelFile(path, engine='openpyxl')
            sheets = {}
            file_changed = False

            for sheet in xl.sheet_names:
                df = xl.parse(sheet)
                df = clean_columns(df)

                col = next((c for c in NAME_COLS if c in df.columns), None)
                if col:
                    for old, new in name_changes.items():
                        mask = df[col] == old
                        if mask.any():
                            df.loc[mask, col] = new
                            cnt = int(mask.sum())
                            print(f"    [{path}] '{sheet}' 시트: {old} → {new} ({cnt}행)")
                            total += cnt
                            file_changed = True

                sheets[sheet] = df

            if file_changed:
                with pd.ExcelWriter(path, engine='openpyxl') as w:
                    for sheet, df in sheets.items():
                        df.to_excel(w, sheet_name=sheet, index=False)

        except Exception as e:
            print(f"  [Excel] {path} 업데이트 실패: {e}")

    return total


def fill_missing_stock_codes(name_to_code: dict, aliases: dict | None = None, target_files: list[str] | None = None) -> int:
    """엑셀의 빈 종목코드를 종목명 기준으로 채움. 기존 종목코드는 덮어쓰지 않음."""
    from openpyxl import load_workbook

    aliases = aliases or {}
    total = 0
    target_files = target_files or EXCEL_FILES

    for path in target_files:
        if not os.path.exists(path):
            continue

        try:
            wb = load_workbook(path)
            file_changed = False

            for ws in wb.worksheets:
                headers = {}
                for cell in ws[1]:
                    header = str(cell.value or '').replace(' ', '').strip()
                    if header:
                        headers[header] = cell.column

                name_col = next((headers[c] for c in NAME_COLS if c in headers), None)
                code_col = next((headers[c] for c in CODE_COLS if c in headers), None)

                if not name_col:
                    continue

                if not code_col:
                    code_col = ws.max_column + 1
                    ws.cell(row=1, column=code_col, value='종목코드')

                sheet_count = 0
                for row_idx in range(2, ws.max_row + 1):
                    name = str(ws.cell(row=row_idx, column=name_col).value or '').strip()
                    if not name:
                        continue

                    current_code = normalize_stock_code(ws.cell(row=row_idx, column=code_col).value)
                    if current_code:
                        continue

                    code = resolve_code_by_name(name, name_to_code, aliases)
                    if not code:
                        continue

                    cell = ws.cell(row=row_idx, column=code_col)
                    cell.value = code
                    cell.number_format = '@'
                    sheet_count += 1
                    total += 1
                    file_changed = True

                if sheet_count:
                    print(f"    [{path}] '{ws.title}' 시트: 종목코드 {sheet_count}행 입력")

            if file_changed:
                wb.save(path)

        except Exception as e:
            print(f"  [Excel] {path} 종목코드 입력 실패: {e}")

    return total


# ── 메인 ───────────────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(description='사명 변경 감지 및 종목코드 자동 입력')
    parser.add_argument(
        '--fill-codes-only',
        action='store_true',
        help='사명 변경/엑셀 종목명 업데이트 없이 빈 종목코드만 채웁니다.',
    )
    parser.add_argument(
        '--offline',
        action='store_true',
        help='FDR/DART 조회 없이 stock_code_map.json과 엑셀에 이미 있는 코드만 사용합니다.',
    )
    parser.add_argument(
        '--target-file',
        action='append',
        help='종목코드를 채울 엑셀 파일을 지정합니다. 여러 번 사용할 수 있으며, 생략하면 관리 대상 전체를 처리합니다.',
    )
    return parser.parse_args()


def main():
    args = parse_args()

    print(f"\n{'='*55}")
    print(f"  사명 변경 자동 업데이트: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'='*55}\n")

    # 1) 현재 KRX 이름→코드 조회 (FDR)
    print("[1] FDR KRX 종목 조회...")
    if args.offline:
        fdr_map, fdr_code_to_name = {}, {}
        print("    --offline 모드: FDR 조회 스킵\n")
    else:
        fdr_map, fdr_code_to_name = get_fdr_maps()
        print(f"    {len(fdr_map)}개 종목 로드\n")

    # 2) DART 코드→현재 이름 조회
    print("[2] DART 법인 정보 조회...")
    if args.offline or args.fill_codes_only:
        dart_map = {}
        print("  [DART] 코드 입력 전용/오프라인 모드 — 스킵")
    else:
        dart_map = get_dart_code_to_name()
    print()

    # 3) 누적 매핑 로드 후 FDR 현재 종목으로 보강
    #    (구 사명 → 코드 기록이 누적되어 있어 이름 변경 후에도 추적 가능)
    print("[3] 코드 매핑 로드...")
    code_map = load_code_map()
    prev_count = len(code_map)
    code_map.update(fdr_map)   # FDR 현재 이름으로 갱신 (신규 상장 포함)
    print(f"    이전 {prev_count}개 → 현재 {len(code_map)}개\n")

    # 4) Excel 전체 종목명/종목코드 수집
    print("[4] Excel 종목명/종목코드 수집...")
    excel_records = get_all_stock_records()
    excel_names = {record['name'] for record in excel_records if record['name']}
    excel_codes = {record['code'] for record in excel_records if record['code']}
    for record in excel_records:
        if record['name'] and record['code']:
            code_map[record['name']] = record['code']
    print(f"    종목명 {len(excel_names)}개, 종목코드 {len(excel_codes)}개 발견\n")

    # 4-0) 같은 종목코드의 과거/현재 이름을 alias로 정리
    print("[4-0] 중복 종목코드 alias 정리...")
    aliases = load_aliases()
    alias_count = sync_aliases_from_duplicate_codes(code_map, fdr_code_to_name, aliases)
    if alias_count:
        save_aliases(aliases)
        print(f"    alias {alias_count}건 추가/갱신\n")
    elif fdr_code_to_name:
        print("    추가할 alias 없음\n")
    else:
        print("    현재 상장명 조회가 없어 스킵\n")

    # 4-1) 빈 종목코드 자동 입력
    print("[4-1] 빈 종목코드 자동 입력...")
    code_fill_count = fill_missing_stock_codes(code_map, aliases, args.target_file)
    if code_fill_count:
        print(f"    총 {code_fill_count}행 종목코드 입력 완료\n")
        excel_records = get_all_stock_records()
        for record in excel_records:
            if record['name'] and record['code']:
                code_map[record['name']] = record['code']
    else:
        print("    입력할 빈 종목코드 없음\n")

    if args.fill_codes_only:
        save_code_map(code_map)
        print(f"[완료] 코드 입력 전용 실행: stock_code_map.json {len(code_map)}개 매핑 저장")
        print(f"\n{'='*55}  완료\n")
        return

    # 5) 사명 변경 감지
    print("[5] 사명 변경 감지...")
    name_changes = {}
    unmapped = []
    current_code_to_name = dict(dart_map)
    current_code_to_name.update(fdr_code_to_name)  # 화면 표시명은 KRX/FDR 이름을 우선 사용

    for record in excel_records:
        old_name = record['name']
        code = record['code'] or code_map.get(old_name)
        if not code:
            if old_name:
                unmapped.append(old_name)
            continue

        current_name = current_code_to_name.get(code)
        if old_name:
            code_map[old_name] = code

        if current_name:
            code_map[current_name] = code

        if old_name and current_name and current_name != old_name:
            name_changes[old_name] = current_name
            print(f"    변경 감지: {old_name} → {current_name}  (코드: {code})")

    if not name_changes:
        print("    변경 없음")

    if unmapped:
        print(f"\n    코드 미매핑 종목 {len(unmapped)}개 (처음 실행 시 정상):")
        for n in unmapped[:10]:
            print(f"      - {n}")
        if len(unmapped) > 10:
            print(f"      ... 외 {len(unmapped)-10}개")
    print()

    # 6) Excel 업데이트
    print("[6] Excel 업데이트...")
    if name_changes:
        total = apply_name_changes(name_changes)
        print(f"    총 {total}행 수정 완료\n")
    else:
        print("    업데이트 없음\n")

    # 7) 매핑 파일 저장
    save_code_map(code_map)
    print(f"[7] 매핑 저장 완료: {len(code_map)}개 종목")

    # 8) 사명 변경 이력(alias) 누적 저장
    if name_changes:
        aliases.update(name_changes)

    # A→B, B→C처럼 누적된 별칭은 항상 최종 사명으로 접어둠
    for old_name in list(aliases.keys()):
        aliases[old_name] = resolve_alias_name(aliases[old_name], aliases)
    save_aliases(aliases)
    print(f"[8] 사명 이력 저장 완료: 누적 {len(aliases)}건")

    print(f"\n{'='*55}  완료\n")


if __name__ == '__main__':
    main()
